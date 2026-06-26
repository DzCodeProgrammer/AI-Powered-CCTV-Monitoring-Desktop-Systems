"""Verify Session 14: attendance Excel export."""

import io
import os
import sys
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
os.chdir(PROJECT_ROOT)
sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import load_dotenv

load_dotenv(PROJECT_ROOT / ".env")


def main() -> int:
    from openpyxl import load_workbook
    from fastapi.testclient import TestClient
    from sqlalchemy import delete

    from app import create_app
    from app.database.connection import SessionLocal, init_db
    from app.models.attendance import Attendance
    from app.services.auth_service import create_admin, get_admin_by_username
    from app.services.export_service import build_attendance_excel, export_attendance_to_excel
    from app.utils.config import get_settings

    get_settings.cache_clear()
    init_db()

    admin_user = "_verify_export_admin_"
    db = SessionLocal()
    try:
        existing = get_admin_by_username(db, admin_user)
        if existing:
            db.delete(existing)
        db.execute(delete(Attendance).where(Attendance.detected_name == "_Verify Export_"))
        db.commit()
        create_admin(db, admin_user, "verify-export-pass")
        db.add(
            Attendance(
                detected_name="_Verify Export_",
                detected_at=datetime.utcnow(),
                camera_source="0",
                status="Recognized",
                confidence=0.91,
            )
        )
        db.commit()
    finally:
        db.close()

    db = SessionLocal()
    try:
        records = db.scalars(
            __import__("sqlalchemy").select(Attendance).where(
                Attendance.detected_name == "_Verify Export_"
            )
        ).all()
        data = build_attendance_excel(list(records))
        workbook = load_workbook(io.BytesIO(data))
        sheet = workbook.active
        if sheet.max_row < 2:
            print("FAIL: Excel should contain header + data row")
            return 1
        if sheet["B2"].value != "_Verify Export_":
            print(f"FAIL: unexpected Excel cell value: {sheet['B2'].value}")
            return 1
        print("Excel build service: OK")

        content, filename = export_attendance_to_excel(db)
        if not filename.endswith(".xlsx"):
            print("FAIL: export filename should end with .xlsx")
            return 1
        print(f"Export filename: {filename}")
    finally:
        db.close()

    client = TestClient(create_app())
    client.post(
        "/login",
        data={"username": admin_user, "password": "verify-export-pass"},
        follow_redirects=False,
    )

    response = client.get("/dashboard/attendance/export")
    if response.status_code != 200:
        print(f"FAIL: export route returned {response.status_code}")
        return 1
    if "spreadsheetml" not in response.headers.get("content-type", ""):
        print("FAIL: export content-type is not Excel")
        return 1
    if len(response.content) < 100:
        print("FAIL: export file too small")
        return 1

    preview = client.get("/dashboard/attendance/export/preview")
    if preview.status_code != 200:
        print(f"FAIL: export preview page returned {preview.status_code}")
        return 1
    if "Export Attendance to Excel" not in preview.text:
        print("FAIL: export preview page content missing")
        return 1

    anon = TestClient(create_app())
    blocked = anon.get("/dashboard/attendance/export", follow_redirects=False)
    if blocked.status_code not in (303, 307):
        print(f"FAIL: unauthenticated export should redirect, got {blocked.status_code}")
        return 1

    print("Session 14 Excel export verification: OK")
    return 0


if __name__ == "__main__":
    sys.exit(main())

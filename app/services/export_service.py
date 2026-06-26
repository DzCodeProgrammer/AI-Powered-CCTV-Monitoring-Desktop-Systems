from __future__ import annotations

"""Export attendance records to Excel (.xlsx)."""

from datetime import date, datetime, time
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.attendance import Attendance


def fetch_attendance_records(
    db: Session,
    start: date | None = None,
    end: date | None = None,
    limit: int = 10_000,
) -> list[Attendance]:
    query = select(Attendance).order_by(Attendance.detected_at.desc()).limit(limit)
    if start is not None:
        query = query.where(Attendance.detected_at >= datetime.combine(start, time.min))
    if end is not None:
        query = query.where(Attendance.detected_at <= datetime.combine(end, time.max))
    return list(db.scalars(query).all())


def build_attendance_excel(records: list[Attendance]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"

    headers = [
        "ID",
        "Person",
        "Detected At",
        "Camera Source",
        "Status",
        "Confidence (%)",
        "User ID",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for record in records:
        confidence_pct = (
            round(record.confidence * 100, 1) if record.confidence is not None else None
        )
        sheet.append(
            [
                record.id,
                record.detected_name,
                record.detected_at.strftime("%Y-%m-%d %H:%M:%S")
                if record.detected_at
                else "",
                record.camera_source,
                record.status,
                confidence_pct,
                record.user_id,
            ]
        )

    for column_index, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in sheet.iter_rows(
            min_row=2,
            min_col=column_index,
            max_col=column_index,
            values_only=True,
        ):
            value = row[0]
            if value is not None:
                max_len = max(max_len, len(str(value)))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max_len + 2, 40)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_attendance_to_excel(
    db: Session,
    start: date | None = None,
    end: date | None = None,
) -> tuple[bytes, str]:
    records = fetch_attendance_records(db, start=start, end=end)
    content = build_attendance_excel(records)
    filename = f"attendance_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return content, filename
